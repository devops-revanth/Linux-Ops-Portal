"""Reports blueprint routes."""
from __future__ import annotations

import csv
import io
import logging

from flask import Response, current_app, render_template, request

from . import reports_bp
from .queries import (
    InfrastructureFilters,
    InventoryReportFilters,
    PatchComplianceFilters,
    SyncReportFilters,
    get_infrastructure_summary,
    get_infrastructure_summary_export,
    get_inventory_report,
    get_inventory_report_export,
    get_patch_compliance_export,
    get_patch_compliance_report,
    get_sync_report,
    get_sync_report_export,
    INFRA_DEFAULT_SORT,
    INFRA_DEFAULT_ORDER,
    INVENTORY_DEFAULT_SORT,
    INVENTORY_DEFAULT_ORDER,
    PATCH_DEFAULT_SORT,
    PATCH_DEFAULT_ORDER,
    SYNC_DEFAULT_SORT,
    SYNC_DEFAULT_ORDER,
)

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────── #
# Export helpers
# ─────────────────────────────────────────────────────────────────────────── #

def _csv_response(rows: list[list], headers: list[str], filename: str) -> Response:
    """Return a streaming CSV download."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _xlsx_response(rows: list[list], headers: list[str], filename: str) -> Response:
    """Return an Excel (.xlsx) download."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.error("openpyxl not installed — falling back to CSV")
        return _csv_response(rows, headers, filename.replace(".xlsx", ".csv"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename.replace(".xlsx", "")[:31]

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _fmt(value) -> str:
    """Safe string formatter for export values."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


# ─────────────────────────────────────────────────────────────────────────── #
# Reports landing page
# ─────────────────────────────────────────────────────────────────────────── #

@reports_bp.route("/reports")
def index():
    return render_template(
        "reports/index.html",
        app_name=current_app.config["APP_NAME"],
        app_version=current_app.config["APP_VERSION"],
    )


# ─────────────────────────────────────────────────────────────────────────── #
# 1. Server Inventory Report
# ─────────────────────────────────────────────────────────────────────────── #

@reports_bp.route("/reports/inventory")
def inventory():
    per_page = current_app.config.get("ITEMS_PER_PAGE", 25)

    search      = request.args.get("q",           "").strip()
    location_id = request.args.get("location_id", type=int)
    env_id      = request.args.get("env_id",      type=int)
    owner_id    = request.args.get("owner_id",    type=int)
    status      = request.args.get("status",      "").strip()
    sort        = request.args.get("sort",  INVENTORY_DEFAULT_SORT)
    order       = request.args.get("order", INVENTORY_DEFAULT_ORDER)
    page        = request.args.get("page",  1, type=int)
    export      = request.args.get("export", "").strip().lower()

    if order not in ("asc", "desc"):
        order = INVENTORY_DEFAULT_ORDER
    if page < 1:
        page = 1

    filters = InventoryReportFilters(
        search=search, location_id=location_id, env_id=env_id,
        owner_id=owner_id, status=status, sort=sort, order=order,
    )

    if export in ("csv", "xlsx"):
        servers = get_inventory_report_export(filters)
        headers = [
            "Hostname", "FQDN", "IP Address", "Environment", "Location", "Owner",
            "OS", "OS Version", "Kernel", "CPU Count", "RAM (GB)", "Status",
            "Last Ansible Sync", "Created At",
        ]
        rows = [
            [
                s.hostname, _fmt(s.fqdn), s.ip_address,
                s.environment.name if s.environment else "",
                s.location.name    if s.location    else "",
                s.owner.name       if s.owner       else "",
                _fmt(s.operating_system), _fmt(s.os_version), _fmt(s.kernel_version),
                _fmt(s.cpu_count), _fmt(s.ram_gb), s.status,
                _fmt(s.last_ansible_sync), _fmt(s.created_at),
            ]
            for s in servers
        ]
        fn = "server_inventory_report"
        if export == "csv":
            return _csv_response(rows, headers, f"{fn}.csv")
        return _xlsx_response(rows, headers, f"{fn}.xlsx")

    report = get_inventory_report(filters, page=page, per_page=per_page)
    return render_template(
        "reports/inventory.html",
        report=report,
        app_name=current_app.config["APP_NAME"],
        app_version=current_app.config["APP_VERSION"],
    )


# ─────────────────────────────────────────────────────────────────────────── #
# 2. Infrastructure Summary
# ─────────────────────────────────────────────────────────────────────────── #

@reports_bp.route("/reports/infrastructure-summary")
def infrastructure_summary():
    per_page = current_app.config.get("ITEMS_PER_PAGE", 25)

    search      = request.args.get("q",           "").strip()
    location_id = request.args.get("location_id", type=int)
    env_id      = request.args.get("env_id",      type=int)
    status      = request.args.get("status",      "").strip()
    sort        = request.args.get("sort",  INFRA_DEFAULT_SORT)
    order       = request.args.get("order", INFRA_DEFAULT_ORDER)
    page        = request.args.get("page",  1, type=int)
    export      = request.args.get("export", "").strip().lower()

    if order not in ("asc", "desc"):
        order = INFRA_DEFAULT_ORDER
    if page < 1:
        page = 1

    filters = InfrastructureFilters(
        search=search, location_id=location_id, env_id=env_id,
        status=status, sort=sort, order=order,
    )

    if export in ("csv", "xlsx"):
        servers = get_infrastructure_summary_export(filters)
        headers = [
            "Hostname", "IP Address", "Location", "Environment",
            "OS", "Owner", "Status", "Last Ansible Sync",
        ]
        rows = [
            [
                s.hostname, s.ip_address,
                s.location.name    if s.location    else "",
                s.environment.name if s.environment else "",
                _fmt(s.operating_system),
                s.owner.name       if s.owner       else "",
                s.status, _fmt(s.last_ansible_sync),
            ]
            for s in servers
        ]
        fn = "infrastructure_summary"
        if export == "csv":
            return _csv_response(rows, headers, f"{fn}.csv")
        return _xlsx_response(rows, headers, f"{fn}.xlsx")

    report = get_infrastructure_summary(filters, page=page, per_page=per_page)
    return render_template(
        "reports/infrastructure_summary.html",
        report=report,
        app_name=current_app.config["APP_NAME"],
        app_version=current_app.config["APP_VERSION"],
    )


# ─────────────────────────────────────────────────────────────────────────── #
# 3. Patch Compliance Report
# ─────────────────────────────────────────────────────────────────────────── #

@reports_bp.route("/reports/patch-compliance")
def patch_compliance():
    per_page = current_app.config.get("ITEMS_PER_PAGE", 25)

    search            = request.args.get("q",                 "").strip()
    location_id       = request.args.get("location_id",       type=int)
    env_id            = request.args.get("env_id",            type=int)
    compliance_status = request.args.get("compliance_status", "").strip()
    sort              = request.args.get("sort",  PATCH_DEFAULT_SORT)
    order             = request.args.get("order", PATCH_DEFAULT_ORDER)
    page              = request.args.get("page",  1, type=int)
    export            = request.args.get("export", "").strip().lower()

    if order not in ("asc", "desc"):
        order = PATCH_DEFAULT_ORDER
    if page < 1:
        page = 1

    filters = PatchComplianceFilters(
        search=search, location_id=location_id, env_id=env_id,
        compliance_status=compliance_status, sort=sort, order=order,
    )

    if export in ("csv", "xlsx"):
        servers = get_patch_compliance_export(filters)
        headers = [
            "Hostname", "Environment", "Location", "OS", "Current Kernel",
            "Previous Kernel", "Compliance Status", "Patch Status",
            "Pending Updates", "Last Patched", "Last Reboot",
            "Last Ansible Sync", "Owner",
        ]
        rows = []
        for s in servers:
            p = s.patching
            rows.append([
                s.hostname,
                s.environment.name if s.environment else "",
                s.location.name    if s.location    else "",
                _fmt(s.operating_system),
                _fmt(p.current_kernel    if p else None),
                _fmt(p.previous_kernel   if p else None),
                p.compliance_status      if p else "unknown",
                p.patch_status           if p else "unknown",
                _fmt(p.pending_updates   if p else None),
                _fmt(p.last_patch_date   if p else None),
                _fmt(p.last_reboot_date  if p else None),
                _fmt(s.last_ansible_sync),
                s.owner.name if s.owner else "",
            ])
        fn = "patch_compliance_report"
        if export == "csv":
            return _csv_response(rows, headers, f"{fn}.csv")
        return _xlsx_response(rows, headers, f"{fn}.xlsx")

    report = get_patch_compliance_report(filters, page=page, per_page=per_page)
    return render_template(
        "reports/patch_compliance.html",
        report=report,
        app_name=current_app.config["APP_NAME"],
        app_version=current_app.config["APP_VERSION"],
    )


# ─────────────────────────────────────────────────────────────────────────── #
# 4. Synchronization Report
# ─────────────────────────────────────────────────────────────────────────── #

@reports_bp.route("/reports/synchronization")
def synchronization():
    per_page = current_app.config.get("ITEMS_PER_PAGE", 25)

    search      = request.args.get("q",           "").strip()
    location_id = request.args.get("location_id", type=int)
    env_id      = request.args.get("env_id",      type=int)
    sync_status = request.args.get("sync_status", "").strip()
    sort        = request.args.get("sort",  SYNC_DEFAULT_SORT)
    order       = request.args.get("order", SYNC_DEFAULT_ORDER)
    page        = request.args.get("page",  1, type=int)
    export      = request.args.get("export", "").strip().lower()

    if sync_status not in ("synced", "never", ""):
        sync_status = ""
    if order not in ("asc", "desc"):
        order = SYNC_DEFAULT_ORDER
    if page < 1:
        page = 1

    filters = SyncReportFilters(
        search=search, location_id=location_id, env_id=env_id,
        sync_status=sync_status, sort=sort, order=order,
    )

    if export in ("csv", "xlsx"):
        servers = get_sync_report_export(filters)
        headers = [
            "Hostname", "IP Address", "Environment", "Location", "OS",
            "Patch Status", "Last Ansible Sync", "Owner",
        ]
        rows = []
        for s in servers:
            p = s.patching
            rows.append([
                s.hostname, s.ip_address,
                s.environment.name if s.environment else "",
                s.location.name    if s.location    else "",
                _fmt(s.operating_system),
                p.patch_status if p else "unknown",
                _fmt(s.last_ansible_sync),
                s.owner.name if s.owner else "",
            ])
        fn = "synchronization_report"
        if export == "csv":
            return _csv_response(rows, headers, f"{fn}.csv")
        return _xlsx_response(rows, headers, f"{fn}.xlsx")

    report = get_sync_report(filters, page=page, per_page=per_page)
    return render_template(
        "reports/synchronization.html",
        report=report,
        app_name=current_app.config["APP_NAME"],
        app_version=current_app.config["APP_VERSION"],
    )

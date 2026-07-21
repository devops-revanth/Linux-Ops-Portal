"""
Audit Log query helpers.

All database reads for the Audit Logs module live here.
Routes stay thin; all filtering, pagination, and export logic is here.
"""
from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone

from ...extensions import db
from ...models.audit_log import AuditLog

logger = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────── #

PER_PAGE = 50

VALID_RESULTS = ("success", "failed")

RETENTION_OPTIONS = (
    (30,  "30 days"),
    (90,  "90 days"),
    (180, "180 days"),
    (365, "365 days"),
)

SORTABLE_COLUMNS = {
    "created_at": AuditLog.created_at,
    "actor":      AuditLog.actor,
    "module":     AuditLog.module,
    "action":     AuditLog.action,
    "target":     AuditLog.target,
    "result":     AuditLog.result,
}


# ── Filter dataclass ──────────────────────────────────────────────────────── #

@dataclass
class AuditFilters:
    date_from:  str = ""
    date_to:    str = ""
    actor:      str = ""
    module:     str = ""
    action:     str = ""
    result:     str = ""
    search:     str = ""
    sort:       str = "created_at"
    order:      str = "desc"
    page:       int = 1

    @classmethod
    def from_request(cls, args: dict) -> "AuditFilters":
        try:
            page = max(1, int(args.get("page", 1)))
        except (ValueError, TypeError):
            page = 1
        sort = args.get("sort", "created_at")
        if sort not in SORTABLE_COLUMNS:
            sort = "created_at"
        order = "asc" if args.get("order", "desc") == "asc" else "desc"
        return cls(
            date_from = args.get("date_from", "").strip(),
            date_to   = args.get("date_to",   "").strip(),
            actor     = args.get("actor",     "").strip(),
            module    = args.get("module",    "").strip(),
            action    = args.get("action",    "").strip(),
            result    = args.get("result",    "").strip(),
            search    = args.get("search",    "").strip(),
            sort      = sort,
            order     = order,
            page      = page,
        )

    def has_active_filters(self) -> bool:
        return any([
            self.date_from, self.date_to, self.actor,
            self.module, self.action, self.result, self.search,
        ])


# ── Query builder ─────────────────────────────────────────────────────────── #

def _apply_filters(query, filters: AuditFilters):
    """Apply all active filters to an AuditLog query."""
    if filters.date_from:
        try:
            dt = datetime.strptime(filters.date_from, "%Y-%m-%d")
            query = query.filter(AuditLog.created_at >= dt)
        except ValueError:
            pass

    if filters.date_to:
        try:
            dt = datetime.strptime(filters.date_to, "%Y-%m-%d") + timedelta(days=1)
            query = query.filter(AuditLog.created_at < dt)
        except ValueError:
            pass

    if filters.actor:
        query = query.filter(
            db.func.lower(AuditLog.actor) == filters.actor.lower()
        )

    if filters.module:
        query = query.filter(
            db.func.lower(AuditLog.module) == filters.module.lower()
        )

    if filters.action:
        query = query.filter(
            AuditLog.action.ilike(f"%{filters.action}%")
        )

    if filters.result and filters.result in VALID_RESULTS:
        query = query.filter(AuditLog.result == filters.result)

    if filters.search:
        term = f"%{filters.search}%"
        query = query.filter(
            db.or_(
                AuditLog.actor.ilike(term),
                AuditLog.action.ilike(term),
                AuditLog.target.ilike(term),
                AuditLog.details.ilike(term),
                AuditLog.ip_address.ilike(term),
            )
        )

    return query


def _apply_sort(query, filters: AuditFilters):
    col = SORTABLE_COLUMNS.get(filters.sort, AuditLog.created_at)
    return query.order_by(col.asc() if filters.order == "asc" else col.desc())


# ── Public query functions ────────────────────────────────────────────────── #

def get_audit_page(filters: AuditFilters):
    """Return a Flask-SQLAlchemy Pagination object for the given filters."""
    try:
        q = _apply_filters(AuditLog.query, filters)
        q = _apply_sort(q, filters)
        return q.paginate(page=filters.page, per_page=PER_PAGE, error_out=False)
    except Exception:
        logger.exception("Failed to query audit log page")
        return None


def get_audit_entry(entry_id: int) -> "AuditLog | None":
    """Return a single audit log entry by ID."""
    try:
        return AuditLog.query.get(entry_id)
    except Exception:
        logger.exception("Failed to fetch audit entry id=%d", entry_id)
        return None


def get_all_for_export(filters: AuditFilters) -> list:
    """Return all matching rows for CSV/Excel export (no pagination limit)."""
    try:
        q = _apply_filters(AuditLog.query, filters)
        q = _apply_sort(q, filters)
        return q.limit(50_000).all()   # hard cap for safety
    except Exception:
        logger.exception("Failed to fetch audit rows for export")
        return []


def get_distinct_actors() -> list[str]:
    """Return sorted list of distinct actors for the filter dropdown."""
    try:
        rows = db.session.query(AuditLog.actor).distinct().order_by(AuditLog.actor).all()
        return [r[0] for r in rows if r[0]]
    except Exception:
        return []


def get_distinct_modules() -> list[str]:
    """Return sorted list of distinct modules for the filter dropdown."""
    try:
        rows = db.session.query(AuditLog.module).distinct().order_by(AuditLog.module).all()
        return [r[0] for r in rows if r[0]]
    except Exception:
        return []


def get_total_count() -> int:
    """Return total number of audit log rows."""
    try:
        return AuditLog.query.count()
    except Exception:
        return 0


# ── Export ────────────────────────────────────────────────────────────────── #

_CSV_HEADERS = [
    "ID", "Timestamp (UTC)", "User", "Module", "Action",
    "Target", "Details", "IP Address", "Auth Source", "Result",
    "User Agent", "Session ID", "Before Values", "After Values",
]


def _row_to_list(entry: AuditLog) -> list:
    return [
        entry.id,
        entry.created_at.strftime("%Y-%m-%d %H:%M:%S") if entry.created_at else "",
        entry.actor,
        entry.module or "",
        entry.action,
        entry.target or "",
        entry.details or "",
        entry.ip_address or "",
        entry.auth_source or "",
        entry.result,
        entry.user_agent or "",
        entry.session_id or "",
        entry.before_values or "",
        entry.after_values or "",
    ]


def build_csv(rows: list) -> str:
    """Return a CSV string from the given AuditLog rows."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_CSV_HEADERS)
    for entry in rows:
        writer.writerow(_row_to_list(entry))
    return output.getvalue()


def build_excel(rows: list) -> bytes:
    """Return an Excel (.xlsx) bytes object from the given AuditLog rows."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"

    # Header row styling
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="F8FAFC")
    ws.append(_CSV_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")

    for entry in rows:
        ws.append(_row_to_list(entry))

    # Auto-size columns (rough estimate)
    col_widths = [6, 20, 14, 14, 30, 25, 30, 16, 12, 10, 30, 20, 20, 20]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Retention / Cleanup ───────────────────────────────────────────────────── #

def purge_old_entries(days: int) -> int:
    """Delete audit log entries older than `days` days.

    Returns the number of rows deleted.  Raises on database error.
    """
    if days < 1:
        raise ValueError("Retention period must be at least 1 day.")
    cutoff = datetime.utcnow() - timedelta(days=days)
    try:
        count = AuditLog.query.filter(AuditLog.created_at < cutoff).delete()
        db.session.commit()
        logger.info("Audit log purge: deleted %d entries older than %d days", count, days)
        return count
    except Exception:
        db.session.rollback()
        logger.exception("Audit log purge failed")
        raise
